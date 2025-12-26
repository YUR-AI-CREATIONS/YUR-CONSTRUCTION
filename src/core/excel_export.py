"""
Excel Export System

Generates perfectly formatted Excel estimates with CSI divisions,
itemized costs, unit pricing, totals, and embedded scopes
"""

from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..utils.csi_divisions import get_division_name, get_all_divisions


class ExcelExporter:
    """
    Creates formatted Excel cost estimates
    """
    
    def __init__(self, output_folder: str = "outputs"):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        
    def create_estimate(self, 
                       project_name: str,
                       agent_results: List[Dict[str, Any]],
                       metadata: Dict[str, Any] = None) -> str:
        """
        Create a comprehensive Excel estimate
        
        Args:
            project_name: Name of the construction project
            agent_results: Results from AI agents
            metadata: Additional project metadata
            
        Returns:
            Path to generated Excel file
        """
        wb = openpyxl.Workbook()
        
        # Create sheets
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        detail_sheet = wb.create_sheet("Detailed Estimate")
        divisions_sheet = wb.create_sheet("CSI Divisions")
        audit_sheet = wb.create_sheet("Audit Trail")
        
        # Populate sheets
        self._create_summary_sheet(summary_sheet, project_name, agent_results, metadata)
        self._create_detail_sheet(detail_sheet, agent_results)
        self._create_divisions_sheet(divisions_sheet)
        self._create_audit_sheet(audit_sheet, agent_results, metadata)
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{project_name.replace(' ', '_')}_{timestamp}.xlsx"
        filepath = self.output_folder / filename
        
        wb.save(filepath)
        return str(filepath)
    
    def _create_summary_sheet(self, 
                            sheet: Any,
                            project_name: str,
                            agent_results: List[Dict[str, Any]],
                            metadata: Dict[str, Any]):
        """Create summary sheet with project overview"""
        # Styling
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(size=16, bold=True)
        
        # Title
        sheet['A1'] = "CONSTRUCTION COST ESTIMATE"
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:F1')
        
        # Project info
        sheet['A3'] = "Project:"
        sheet['B3'] = project_name
        sheet['A4'] = "Date:"
        sheet['B4'] = datetime.now().strftime("%Y-%m-%d")
        sheet['A5'] = "Prepared By:"
        sheet['B5'] = "BID-ZONE AI Estimating Platform"
        
        # Cost summary header
        row = 7
        headers = ['CSI Division', 'Description', 'Quantity', 'Unit', 'Unit Price', 'Total Cost']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Aggregate costs by division
        division_costs = {}
        for result in agent_results:
            data = result.get('data', {})
            division = data.get('csi_division', '99')
            items = data.get('items', [])
            
            if division not in division_costs:
                division_costs[division] = {
                    'description': get_division_name(division),
                    'items': [],
                    'total': 0
                }
            
            for item in items:
                division_costs[division]['items'].append(item)
                division_costs[division]['total'] += item.get('total', 0)
        
        # Write division summaries
        row += 1
        grand_total = 0
        for division in sorted(division_costs.keys()):
            data = division_costs[division]
            sheet.cell(row=row, column=1).value = f"DIV {division}"
            sheet.cell(row=row, column=2).value = data['description']
            sheet.cell(row=row, column=6).value = data['total']
            sheet.cell(row=row, column=6).number_format = '$#,##0.00'
            grand_total += data['total']
            row += 1
        
        # Grand total
        row += 1
        sheet.cell(row=row, column=5).value = "GRAND TOTAL:"
        sheet.cell(row=row, column=5).font = Font(bold=True)
        sheet.cell(row=row, column=6).value = grand_total
        sheet.cell(row=row, column=6).number_format = '$#,##0.00'
        sheet.cell(row=row, column=6).font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in range(1, 7):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_detail_sheet(self, sheet: Any, agent_results: List[Dict[str, Any]]):
        """Create detailed estimate sheet with line items"""
        # Styling
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        division_font = Font(size=11, bold=True)
        division_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Headers
        row = 1
        headers = ['Item #', 'CSI Div', 'Description', 'Scope', 'Quantity', 'Unit', 'Unit Price', 'Total', 'Agent']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Group by division
        division_groups = {}
        for result in agent_results:
            data = result.get('data', {})
            division = data.get('csi_division', '99')
            
            if division not in division_groups:
                division_groups[division] = []
            
            division_groups[division].append(result)
        
        # Write items
        row = 2
        item_num = 1
        
        for division in sorted(division_groups.keys()):
            # Division header
            division_name = get_division_name(division)
            sheet.cell(row=row, column=1).value = f"DIVISION {division}"
            sheet.merge_cells(f'A{row}:I{row}')
            cell = sheet.cell(row=row, column=1)
            cell.font = division_font
            cell.fill = division_fill
            row += 1
            
            # Items in this division
            for result in division_groups[division]:
                data = result.get('data', {})
                scope = data.get('scope', '')
                agent_id = result.get('agent_id', 'unknown')
                
                for item in data.get('items', []):
                    sheet.cell(row=row, column=1).value = item_num
                    sheet.cell(row=row, column=2).value = division
                    sheet.cell(row=row, column=3).value = item.get('description', '')
                    sheet.cell(row=row, column=4).value = scope
                    sheet.cell(row=row, column=5).value = item.get('quantity', 0)
                    sheet.cell(row=row, column=6).value = item.get('unit', '')
                    sheet.cell(row=row, column=7).value = item.get('unit_price', 0)
                    sheet.cell(row=row, column=7).number_format = '$#,##0.00'
                    sheet.cell(row=row, column=8).value = item.get('total', 0)
                    sheet.cell(row=row, column=8).number_format = '$#,##0.00'
                    sheet.cell(row=row, column=9).value = agent_id
                    
                    item_num += 1
                    row += 1
            
            row += 1  # Blank row between divisions
        
        # Auto-adjust column widths
        column_widths = [8, 8, 30, 25, 10, 8, 12, 12, 15]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
    
    def _create_divisions_sheet(self, sheet: Any):
        """Create CSI divisions reference sheet"""
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        sheet['A1'] = "CSI Division"
        sheet['B1'] = "Name"
        sheet['C1'] = "Description"
        
        for cell in ['A1', 'B1', 'C1']:
            sheet[cell].font = header_font
            sheet[cell].fill = header_fill
        
        divisions = get_all_divisions()
        row = 2
        for code, info in sorted(divisions.items()):
            sheet.cell(row=row, column=1).value = code
            sheet.cell(row=row, column=2).value = info['name']
            sheet.cell(row=row, column=3).value = info['description']
            row += 1
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 50
    
    def _create_audit_sheet(self, 
                           sheet: Any,
                           agent_results: List[Dict[str, Any]],
                           metadata: Dict[str, Any]):
        """Create audit trail sheet"""
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = ['Timestamp', 'Agent ID', 'Specialty', 'Chunk ID', 'Status', 'Items Extracted']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Audit entries
        row = 2
        for result in agent_results:
            sheet.cell(row=row, column=1).value = result.get('timestamp', '')
            sheet.cell(row=row, column=2).value = result.get('agent_id', '')
            sheet.cell(row=row, column=3).value = result.get('specialty', '')
            sheet.cell(row=row, column=4).value = result.get('chunk_id', '')
            sheet.cell(row=row, column=5).value = result.get('status', '')
            
            data = result.get('data', {})
            items_count = len(data.get('items', []))
            sheet.cell(row=row, column=6).value = items_count
            
            row += 1
        
        # Metadata
        if metadata:
            row += 2
            sheet.cell(row=row, column=1).value = "Processing Metadata"
            sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            for key, value in metadata.items():
                sheet.cell(row=row, column=1).value = str(key)
                sheet.cell(row=row, column=2).value = str(value)
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 7):
            sheet.column_dimensions[get_column_letter(col)].width = 20
